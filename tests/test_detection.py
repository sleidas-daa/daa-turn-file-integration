"""Tests for template detection."""
import pytest
from pathlib import Path
from converter.detection import detect_template


class TestDetectTemplate:
    def test_ryanair_by_filename(self, tmp_path):
        p = tmp_path / "FR S26 AC - ryanair adjusted.csv"
        p.write_text("AAl,AFn,DAl,DFn,Frq,Eff,Dsc,Ovn\nFR,11,FR,10,...4...,2026-04-02,2026-10-22,\n")
        template, conf = detect_template(str(p))
        assert template == "ryanair"
        assert conf >= 0.4

    def test_emerald_by_filename(self, tmp_path):
        import openpyxl
        p = tmp_path / "DUB plot - emerald airlines.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["DAY", "EAI AT7 FAT"])
        wb.save(p)
        template, conf = detect_template(str(p))
        assert template == "emerald"
        assert conf >= 0.4

    def test_aer_lingus_by_extension(self, tmp_path):
        p = tmp_path / "EI_schedule.txt"
        p.write_text("3 EI 3220 29MAR2624OCT261234567DUB0600  GLA0715  \n")
        template, conf = detect_template(str(p))
        assert template == "aer_lingus"
        assert conf >= 0.4

    def test_ryanair_by_column_headers(self, tmp_path):
        import openpyxl
        p = tmp_path / "unknown_schedule.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["AAl", "AFn", "DAl", "DFn", "Frq", "Eff", "Dsc", "Ovn", "Apt"])
        ws.append(["FR", 11, "FR", 10, "...4...", "2026-04-02", "2026-10-22", None, "DUB"])
        wb.save(p)
        template, conf = detect_template(str(p))
        assert template == "ryanair"
        assert conf >= 0.5

    def test_unknown_file(self, tmp_path):
        p = tmp_path / "random_data.xlsx"
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ColA", "ColB", "ColC"])
        ws.append([1, 2, 3])
        wb.save(p)
        template, conf = detect_template(str(p))
        assert template == "unknown" or conf < 0.5

    def test_real_ryanair_file(self):
        p = r"C:\Users\Igi20\Downloads\S26 DUB TR - ryanair.xlsx"
        if not Path(p).exists():
            pytest.skip("Real file not available")
        template, conf = detect_template(p)
        assert template == "ryanair"
        assert conf >= 0.5

    def test_real_emerald_file(self):
        p = r"C:\Users\Igi20\Downloads\DUB plot 27012026 - emerald airlines.xlsx"
        if not Path(p).exists():
            pytest.skip("Real file not available")
        template, conf = detect_template(p)
        assert template == "emerald"
        assert conf >= 0.4
