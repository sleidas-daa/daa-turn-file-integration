"""
Shared fixtures and session-level hooks for the test suite.

ENABLE_TESTS
------------
If config.ENABLE_TESTS is False every collected test is skipped.  This is a
last-resort escape hatch — leaving it False hides regressions, so the default
is True.

Fixtures
--------
The synthetic fixtures below build minimal but structurally correct input files
so that parser tests do not depend on any specific real-world schedule file.
Tests that use real files (emerald_plot_cw17, etc.) are guarded with
pytest.skip so they are silently omitted when the file is absent.
"""
import sys
from datetime import date, time
from pathlib import Path

import openpyxl
import pytest

# Allow tests to import the src package without installing it
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

# Import config AFTER adjusting sys.path
from converter import config  # noqa: E402

FIXTURES_DIR = Path(__file__).parent / "fixtures"
EMERALD_PLOT_CW17 = FIXTURES_DIR / "Plot_DUB_CW17Nov25.xlsx"


# ---------------------------------------------------------------------------
# Session-level hook: skip everything when testing is disabled in config
# ---------------------------------------------------------------------------

def pytest_collection_modifyitems(items):
    """Skip all tests when config.ENABLE_TESTS is False."""
    if not config.ENABLE_TESTS:
        skip_marker = pytest.mark.skip(
            reason="Tests disabled via config.ENABLE_TESTS = False"
        )
        for item in items:
            item.add_marker(skip_marker)


# ---------------------------------------------------------------------------
# Ryanair fixture: minimal valid xlsx
# ---------------------------------------------------------------------------

@pytest.fixture()
def ryanair_xlsx(tmp_path) -> Path:
    """Minimal Ryanair-style xlsx with 3 valid DUB turns + 1 bank + 1 non-DUB."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "Apt", "Acd", "Rcd", "AAl", "AFn", "ASx", "DAl", "DFn",
        "DSx", "Tst", "Eff", "Dsc", "Frq", "Ovn",
    ]
    ws.append(headers)
    # Row 1: Thursday turn, no overnight
    ws.append(["DUB", None, None, "FR", 11, None, "FR", 10, None, "N",
                date(2026, 4, 2), date(2026, 10, 22), "...4...", None])
    # Row 2: Monday turn, no overnight
    ws.append(["DUB", None, None, "FR", 11, None, "FR", 10, None, "N",
                date(2026, 3, 30), date(2026, 10, 19), "1......", None])
    # Row 3: Wednesday overnight turn
    ws.append(["DUB", None, None, "FR", 17, None, "FR", 94, None, "N",
                date(2026, 4, 1), date(2026, 10, 21), "..3....", 1])
    # Row 4: bank row (no arrival data)
    ws.append(["DUB", None, None, None, None, None, "FR", 10, None, "N",
                date(2026, 3, 29), date(2026, 3, 29), "......7", None])
    # Row 5: non-hub airport — excluded by hub filter
    ws.append(["ORK", None, None, "FR", 20, None, "FR", 21, None, "N",
                date(2026, 4, 1), date(2026, 10, 21), "..3....", None])
    p = tmp_path / "ryanair_test.xlsx"
    wb.save(p)
    return p


# ---------------------------------------------------------------------------
# Emerald fixtures: minimal valid xlsx with one aircraft column
# ---------------------------------------------------------------------------

@pytest.fixture()
def emerald_xlsx(tmp_path) -> Path:
    """
    Minimal Emerald DUB plot with one aircraft, Monday and Tuesday sections.

    Turns produced (at DUB):
    • Mon: EI3221 (GLA→DUB) + EI3222 (DUB→EDI) — same-day, overnight=0
    • Mon: EI3223 (EDI→DUB) + EI3224 (DUB→GLA) — cross-midnight, overnight=1
    • Tue: EI3225 (GLA→DUB) + EI3226 (DUB→EDI) — same-day, overnight=0
    • Tue→Mon wrap: EI3227 (EDI→DUB Tue) + EI3220 (DUB→GLA Mon) — 6 nights
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    for _ in range(6):
        ws.append([None] * 13)   # rows 1-6: blank (layout preamble)
    # Row 7: header row — parser detects 'DAY' in column A
    ws.append(["DAY", "EAI AT7 FAT", None, None, None, None, None])
    ws.append([None] * 7)        # row 8: blank spacer

    # --- Monday (day 1) ---
    ws.append(["Monday 13JUL26",
                "EI 3220", "DUB", time(5, 50), "GLA", time(7, 5), None])
    ws.append([None,
                "EI 3221", "GLA", time(7, 35), "DUB", time(8, 50), None])
    ws.append([None,
                "EI 3222", "DUB", time(9, 20), "EDI", time(10, 40), None])
    ws.append([None,
                "EI 3223", "EDI", time(11, 20), "DUB", time(12, 40), None])
    ws.append([None] * 7)

    # --- Tuesday (day 2) ---
    ws.append(["Tuesday 14JUL26",
                "EI 3224", "DUB", time(13, 0), "GLA", time(14, 15), None])
    ws.append([None,
                "EI 3225", "GLA", time(14, 55), "DUB", time(16, 10), None])
    ws.append([None,
                "EI 3226", "DUB", time(17, 0), "EDI", time(18, 10), None])
    ws.append([None,
                "EI 3227", "EDI", time(18, 40), "DUB", time(19, 55), None])

    p = tmp_path / "emerald_test.xlsx"
    wb.save(p)
    return p


@pytest.fixture()
def emerald_xlsx_compact_days(tmp_path) -> Path:
    """Emerald fixture with compact day headers that omit the space before the date."""
    wb = openpyxl.Workbook()
    ws = wb.active

    for _ in range(6):
        ws.append([None] * 13)
    ws.append(["DAY", "EAI AT7 FAT", None, None, None, None, None])
    ws.append([None] * 7)

    ws.append(["Monday13JUL26",
                "EI 3220", "DUB", time(5, 50), "GLA", time(7, 5), None])
    ws.append([None,
                "EI 3221", "GLA", time(7, 35), "DUB", time(8, 50), None])
    ws.append([None,
                "EI 3223", "EDI", time(11, 20), "DUB", time(12, 40), None])
    ws.append([None] * 7)

    ws.append(["Tuesday14JUL26",
                "EI 3224", "DUB", time(13, 0), "GLA", time(14, 15), None])
    ws.append([None,
                "EI 3225", "GLA", time(14, 55), "DUB", time(16, 10), None])

    p = tmp_path / "emerald_compact_days.xlsx"
    wb.save(p)
    return p


@pytest.fixture()
def emerald_plot_cw17() -> Path:
    """Real Emerald DUB plot (CW17 Nov 2025): header on row 5, 10 aircraft columns."""
    if not EMERALD_PLOT_CW17.exists():
        pytest.skip(f"Real fixture not found: {EMERALD_PLOT_CW17}")
    return EMERALD_PLOT_CW17


# ---------------------------------------------------------------------------
# Aer Lingus SSIM fixture
# ---------------------------------------------------------------------------

def _ssim_line(airline: str, fnum: str, eff: str, dsc: str, days: str,
               dep: str, arr: str,
               dep_time: str = "0000", arr_time: str = "0000") -> str:
    """
    Build one SSIM Type-3 line matching the extended EI field layout.

    Field layout (0-indexed positions):
      [0-1]   '3 '         record-type + service indicator
      [2-4]   airline      3-char IATA code (space-padded)
      [5-8]   fnum         4-char flight number (space-padded, right-justified)
      [9-13]  '0101J'      itinerary-var / leg / service-type fields
      [14-20] eff          effective date DDMMMYY
      [21-27] dsc          discontinue date DDMMMYY
      [28-34] days         7-char ops-day mask
      [35]    ' '          frequency rate (space)
      [36-38] dep          departure airport
      [39-42] dep_time     departure time local HHMM
      [43-46] dep_time     departure time UTC   HHMM (repeated for simplicity)
      [47-51] '+0000'      UTC variation at departure
      [52-53] '  '         pad
      [54-56] arr          arrival airport
      [57-60] arr_time     arrival time local HHMM
      [61-64] arr_time     arrival time UTC   HHMM (repeated for simplicity)
      [65-69] '+0000'      UTC variation at arrival
      [70-71] '  '         pad / end of core fields
    """
    fnum_padded = f"{fnum:>4}"   # right-justify in 4-char field
    return (
        f"3 {airline:<3}{fnum_padded}0101J"   # positions 0-13
        f"{eff}{dsc}{days} "                   # positions 14-35
        f"{dep}{dep_time}{dep_time}+0000  "   # positions 36-53
        f"{arr}{arr_time}{arr_time}+0000  \n" # positions 54+
    )


@pytest.fixture()
def aer_lingus_txt(tmp_path) -> Path:
    """
    Minimal SSIM TXT fixture that produces exactly one DUB turn.

    EI3221 (GLA→DUB, arrives 08:50) followed by EI3222 (DUB→LHR, departs 09:20).
    The time-based matcher picks EI3222 as the first DUB departure after 08:50.
    EI3220 (DUB→GLA, departs 05:50) is earlier than the arrival so is not matched.
    EI9999 (DUB→BRS) operates on different days so is a separate turn candidate.
    """
    content = (
        "1IATA SSIM\n"
        # EI3220: DUB→GLA (outbound before EI3221 arrives — not a match target)
        + _ssim_line("EI", "3220", "29MAR26", "24OCT26", "1234567",
                     "DUB", "GLA", dep_time="0550", arr_time="0705")
        # EI3221: GLA→DUB (inbound, arrives DUB 08:50)
        + _ssim_line("EI", "3221", "29MAR26", "24OCT26", "1234567",
                     "GLA", "DUB", dep_time="0735", arr_time="0850")
        # EI3222: DUB→LHR (first DUB departure after 08:50 → matched to EI3221)
        + _ssim_line("EI", "3222", "29MAR26", "24OCT26", "1234567",
                     "DUB", "LHR", dep_time="0920", arr_time="1040")
        # EI9999: DUB→BRS (different days mask — won't match EI3221 on 1234567)
        + _ssim_line("EI", "9999", "29MAR26", "24OCT26", "123456 ",
                     "DUB", "BRS", dep_time="1000", arr_time="1100")
        + "9END\n"
    )
    p = tmp_path / "aerlingus_test.txt"
    p.write_text(content, encoding="utf-8")
    return p
